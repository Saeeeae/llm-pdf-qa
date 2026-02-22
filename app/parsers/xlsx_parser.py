import logging
from openpyxl import load_workbook

from app.parsers.base import BaseParser, ParseResult, ParsedPage

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def parse(self, file_path: str) -> ParseResult:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        pages = []
        all_texts = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cell_values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in cell_values):
                    rows.append(" | ".join(cell_values))

            if rows:
                sheet_text = f"## Sheet: {sheet_name}\n\n" + "\n".join(rows)
                pages.append(ParsedPage(page_num=sheet_idx + 1, text=sheet_text))
                all_texts.append(sheet_text)

        wb.close()

        full_text = "\n\n".join(all_texts)

        return ParseResult(
            pages=pages,
            raw_text=full_text,
            total_pages=len(pages),
            metadata={"parser": "openpyxl", "source": file_path},
        )
